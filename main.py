from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Dict, Any
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from io import BytesIO
from fastapi.responses import StreamingResponse, JSONResponse

app = FastAPI()

# Configuración de CORS para permitir todos los orígenes
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],            # Permitir todos los orígenes
    allow_credentials=True,         # Permitir el envío de cookies y credenciales
    allow_methods=["*"],            # Permitir todos los métodos HTTP (GET, POST, etc.)
    allow_headers=["*"],            # Permitir todos los encabezados
)

# Definición de los modelos Pydantic para validar el JSON de entrada
class SheetData(BaseModel):
    hoja: str = Field(..., example="Reporte1")
    title: str = Field(..., example="Reporte de Ventas")
    column_widths: Dict[str, float] = Field(..., example={"Producto": 20, "Cantidad": 15, "Precio": 15})
    data: List[Dict[str, Any]] = Field(..., example=[
        {"Producto": "Manzanas", "Cantidad": 50, "Precio": 1.5},
        {"Producto": "Naranjas", "Cantidad": 30, "Precio": 2.0}
    ])

class ExcelRequest(BaseModel):
    hojas: List[SheetData]

@app.post("/crear-excel")
def crear_excel(request: ExcelRequest):
    try:
        # Crear un nuevo libro de Excel
        wb = Workbook()
        
        # Si no hay hojas en la solicitud, devolver error
        if not request.hojas:
            raise HTTPException(status_code=400, detail="La lista de hojas está vacía.")
        
        for idx, sheet in enumerate(request.hojas):
            # Para la primera hoja, usar la hoja activa
            if idx == 0:
                ws = wb.active
                ws.title = sheet.hoja
            else:
                ws = wb.create_sheet(title=sheet.hoja)
            
            # Agregar el título
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sheet.column_widths))
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = sheet.title
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Alineación vertical centrada
            
            # Agregar los encabezados de columna
            headers = list(sheet.column_widths.keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Alineación vertical centrada
                # Establecer el ancho de la columna
                ws.column_dimensions[get_column_letter(col_num)].width = sheet.column_widths[header]
            
            # Agregar los datos
            for row_num, row_data in enumerate(sheet.data, start=3):
                for col_num, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    # Alineación vertical centrada
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Habilitar el filtrado automático en las columnas
            # Determinar el rango de la tabla (desde la primera columna hasta la última, y desde la fila de encabezados hasta la última fila de datos)
            last_row = ws.max_row
            last_col = len(headers)
            start_cell = f"A2"
            end_cell = f"{get_column_letter(last_col)}{last_row}"
            ws.auto_filter.ref = f"{start_cell}:{end_cell}"
            
            # Opcional: Ajustar la altura de las filas automáticamente
            # Nota: OpenPyXL no soporta auto ajuste de altura de filas, pero Excel lo hace al abrir el archivo
            # Por lo tanto, no es necesario establecer la altura manualmente

        # Guardar el libro en un buffer en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Preparar la respuesta como un archivo de Excel
        headers = {
            'Content-Disposition': 'attachment; filename=archivo_multisheets.xlsx'
        }
        return StreamingResponse(
            buffer,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/excel-a-json")
async def excel_a_json(file: UploadFile = File(...)):
    """
    Endpoint inverso que recibe un archivo Excel y devuelve un JSON
    con la misma estructura que la clase ExcelRequest.
    """
    try:
        # Verificar que el contenido sea un archivo de Excel
        if not file.filename.endswith(('.xlsx', '.xlsm')):
            raise HTTPException(status_code=400, detail="El archivo proporcionado no es un Excel válido.")

        # Cargar el archivo en memoria
        contents = await file.read()
        
        # Cargar el libro de Excel con openpyxl
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
        
        # Estructura final a retornar
        hojas_data = []

        for ws in wb.worksheets:
            # 1) Nombre de la hoja
            hoja_nombre = ws.title
            
            # 2) Título (asumimos que está en la celda A1, fila 1, col 1)
            title_cell = ws.cell(row=1, column=1).value
            # Evitar error si la celda está vacía
            title_value = title_cell if title_cell else ""
            
            # 3) Leer los encabezados (fila 2)
            headers = []
            col = 1
            while True:
                cell_value = ws.cell(row=2, column=col).value
                if cell_value is None:
                    # Se asume que cuando ya no hay encabezado, se termina
                    break
                headers.append(cell_value)
                col += 1
            
            # 4) Reconstruir column_widths 
            #    (mapeando cada "header" a su ancho de columna, si existe)
            column_widths = {}
            for idx, header in enumerate(headers, start=1):
                column_letter = get_column_letter(idx)
                # Puede retornar None si la columna no tiene ancho asignado explícitamente
                col_width = ws.column_dimensions[column_letter].width
                # Si no hubiera ancho establecido, coloca un valor por defecto
                column_widths[header] = col_width if col_width else 10.0
            
            # 5) Leer los datos desde la fila 3 en adelante
            data = []
            row_num = 3
            while True:
                # Detectar si la fila ya no tiene datos
                # Se hace revisando si en la columna 1 ya no hay nada
                first_col_value = ws.cell(row=row_num, column=1).value
                if first_col_value is None:
                    # Podrías refinar la condición para datos esparcidos, 
                    # pero se asume que si la primera columna está vacía, se acabaron los datos
                    break
                
                # Construir el diccionario para la fila
                row_dict = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell_value = ws.cell(row=row_num, column=col_idx).value
                    row_dict[header] = cell_value
                data.append(row_dict)
                row_num += 1
            
            # 6) Armar la estructura tipo SheetData
            sheet_info = {
                "hoja": hoja_nombre,
                "title": title_value,
                "column_widths": column_widths,
                "data": data
            }
            
            hojas_data.append(sheet_info)
        
        # 7) Retornar la respuesta en la misma estructura que ExcelRequest
        response = {
            "hojas": hojas_data
        }
        
        return JSONResponse(content=response, status_code=200)
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))